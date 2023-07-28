from dotenv import load_dotenv
import pandas as pd
import requests
import openpyxl
import shutil
import sys
import os

load_dotenv()

#function to create folder

def createFolder(nameOfFolder):

    if not os.path.exists(nameOfFolder):

        os.mkdir(nameOfFolder)

        folderName = nameOfFolder

        return folderName
 
    while True:

        confirmAction = input("Folder name ({}) already exist, do you want to overwrite [y/n]: ".format(nameOfFolder))

        if confirmAction == 'y' or confirmAction == 'Y':

            shutil.rmtree(nameOfFolder)

            os.mkdir(nameOfFolder)

            folderName = nameOfFolder

            break

            
        elif confirmAction == 'n' or confirmAction == 'N':

            newdir = input("Input new folder name: ")

            folderName = createFolder(newdir)

            break

        else:

            print('\nWrong value entered')
    
    return folderName

# function to send message via telegram bot

def sendToTelegram(document, caption):

    apiToken = os.getenv('botToken')

    chatID = os.getenv('chatID')

    payload = {'chat_id': chatID, 'caption': caption}
    
    apiURL = f'https://api.telegram.org/bot{apiToken}/sendDocument'

    try:

        with open(document, 'rb') as file:

            fileContent = file.read()

        if not fileContent:

            apiURL = f'https://api.telegram.org/bot{apiToken}/sendMessage'

            if document == className + ' termination List.txt':

                text = "No student in {} is due for termination".format(className)

            else:

                text = "No student in {} is due for probation".format(className)

            payload = {'chat_id': chatID, 'text': text}

            response = requests.post(apiURL, data=payload)

            if response.status_code == 200:

                print("Message sent successfully!")

            else:

                print(f"Failed to send message: {response.text}")

            return

        files = {'document': (os.path.basename(document), fileContent)}

        response = requests.post(apiURL, data=payload, files=files)

        if response.status_code == 200:

            print("Document sent successfully!")

        else:

            print(f"Failed to send document: {response.text}")
            
    except Exception as e:

        print(e)

#specify result sheet

resultSheet = "emyety.xlsx"

# allow user to select programe name

print ("1. EMY \n2. ETY")

try:

    programe = int(input("Kindly input which programe to generate result for: "))

    if programe != 1 and programe != 2:

        print('wrong value entered')

        os.execv(sys.executable, ['python'] + sys.argv)
    
    if programe == 1:

        programe = 'EMY'

        resultTemplate = 'emyTemplate.xlsx'

    else:

        programe = "ETY"

        resultTemplate = 'etyTemplate.xlsx'

except ValueError:

    print('Please select from the above option')

    os.execv(sys.executable, ['python'] + sys.argv)

# allow user to select class number

while True:

    try:

        classNumber = int(input("Please input class number only, excluding 'C': "))

        className = programe+"-C"+str(classNumber)

        resultDb = openpyxl.load_workbook(resultSheet)
        classResult = resultDb.sheetnames
        
        if className not in classResult:

            print("\n" + className + " does not exist in excel database\nKindly provide correct class details or update excel database\n")

            os.execv(sys.executable, ['python'] + sys.argv)

        while True:

            confirmSelection = input("Do you want to generate result sheet for every student in "+ className + " (Y/N): ")

            if confirmSelection == "y" or confirmSelection == "Y":

                folder = className + " individual compiled result"
                
                path = createFolder(folder)

                break

            elif confirmSelection == "n" or confirmSelection == "N":

                os.execv(sys.executable, ['python'] + sys.argv)

            else:

                print ("wrong value entered")

        break

    except ValueError:

        print("Please input a number\n")

# Input current number of courses
# selected class that code is working on

currentClass = resultDb[className]

# count number of courses 

# Specify the row number to start the count

row_number_to_search = 3
courseCounted = 0
startColumn = 0

# get first column that course count begins

for cell in currentClass[row_number_to_search]:

    cell_value = cell.value

    if cell_value is not None:
        
        startColumn = cell.column

        break

# start count to know how many courses

for cell in currentClass.iter_cols(min_col=startColumn, min_row=row_number_to_search, max_row=row_number_to_search):
    
    cell_value = cell[0].value

    if cell_value != "AVERAGE":

        courseCounted += 1

    else:
        
        totalCourse = courseCounted + 1 # including column for average(count start from index zero)
        averageValue = courseCounted

        break

# iteration to assign how many student are available while keeping max at 24

# max number of student in class

maxStudent = 24 # that is the maximum number of student in a class
startRow = 7 # from main db names majorly starts from column 7
cellCount = 0
startColumn = 1
numberOfStudent = 0

for cell in currentClass.iter_rows(min_col=startColumn, min_row=startRow):

    cell_value = cell[0].value

    if cell_value is not None and cellCount <= maxStudent:
    
        numberOfStudent += 1
        cellCount += 1
    
    else:

        break

#course index for db

courseRowDb = 5
courseColumnDb = 6

#course index for template

courseRowTemplate = 11
courseColumnTemplate = 3

# other variable

scoreList = []
currentCourseColumn = 11
trackFailure = 0
probation = []
termination = []
courseNameDb = []
courseNameTemplate = []

#load work sheet        

template = openpyxl.load_workbook(resultTemplate)
templateWorksheet = template.worksheets[0]

data = pd.read_excel(resultSheet, sheet_name=className, skiprows=6, header=None, nrows=numberOfStudent) #skip rows is 6 because after that row student names which is the data we need starts counting

# extract all course name from db and template

for i in range(courseCounted):

    dbCell = currentClass.cell(row=courseRowDb, column=courseColumnDb).value
    courseNameDb.append(dbCell)

    courseColumnDb += 1

for i in range(courseCounted):

    templateCell = templateWorksheet.cell(row=courseRowTemplate, column=courseColumnTemplate).value
    courseNameTemplate.append(templateCell)

    courseRowTemplate += 1
                
# iterate through templateWorksheet and manipulate data

for _, row in data.iterrows():

    lastName = row.iloc[1]
    firstName = row.iloc[2]
    middleName = row.iloc[3]
    
    # Check if any of the names is NaN, if yes, set them to empty strings

    if pd.isna(firstName):

        firstName = ""
    
    if pd.isna(middleName):
    
        middleName = ""

    # ommited last name(every individual would surely have a last name) to keep track of empty result that would be called nan
    
    score = row.iloc[5:].values

    score = pd.Series(score).fillna('NA').values

    for scores in score[:totalCourse]:

        scoreList.append(scores)

    average = scoreList[averageValue]

    # check for failures

    for scores in scoreList[:averageValue]:

        try:

            if int(scores) < 60:

                trackFailure += 1
        
        except:

            pass

    # Create a dictionary to store the mapping of course names and their corresponding scores
    course_score_mapping = dict(zip(courseNameDb, score[:totalCourse]))

    # Iterate through the course names in the template
    for coursesTemplate in courseNameTemplate:

        # Check if the course name exists in the mapping
        if coursesTemplate in course_score_mapping:

            # Get the score from the mapping for the current course name
            score_for_course = course_score_mapping[coursesTemplate]

        else:

            # If the course name is not found in the mapping, set the score to 'NA'
            score_for_course = ' '
        # Write the score to the corresponding cell in the template

        templateWorksheet.cell(row=currentCourseColumn, column=4, value=score_for_course)
        currentCourseColumn += 1

    name = str(lastName) + " " + str(firstName) + " " + str(middleName)

    templateWorksheet.cell(row=6, column=3, value=name)
    templateWorksheet.cell(row=6, column=5, value=average)
    templateWorksheet.cell(row=4, column=5, value=className)

    # check number of failed courses

    if trackFailure == 2:

        probation.append(name)

    if trackFailure > 2:

        termination.append(name)

    # clear variable memory

    trackFailure = 0
    scoreList = []
    currentCourseColumn = 11

    #save file

    try:

        template.save('{}/{}.xlsx'.format(path,name))
    
    except:

        pass

# loop to delete null report file created cause of incomplete class

for filename in os.listdir(path):
        
    if "nan" in filename:

        file_path = os.path.join(path, filename)

        os.remove(file_path)

# create file to store probation and termination list

probationFile = open(className + " probation List.txt", "x")
terminationFile = open(className + " termination List.txt", "x")

# generate student on probation

for watchlist in probation:
    
    probationFile.write(watchlist + '\n')

# generate student on termination

for watchlist in termination:
    
    terminationFile.write(watchlist + '\n')

# Save and close files before sending

probationFile.close()
terminationFile.close()

# zip folder containing results

shutil.make_archive(className + " generated result", 'zip', path)
shutil.rmtree(path)

# send generated data to telegram

caption = "This is a TXT file containing list of student in {} due for probation".format(className)

sendToTelegram(className + " probation List.txt", caption)

caption = "This is a TXT file containing list of student in {} due for termination".format(className)

sendToTelegram(className + " termination List.txt", caption)

caption = "This is a zip file containing result for every student in {}".format(className)

sendToTelegram(className + " generated result.zip", caption)

#delete file after sending

os.remove(className + " generated result.zip")
os.remove(className + " termination List.txt")
os.remove(className + " probation List.txt")