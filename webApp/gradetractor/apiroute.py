from flask_jwt_extended import create_access_token, get_jwt_identity, jwt_required, decode_token, jwt_manager, create_refresh_token, set_access_cookies
from werkzeug.security import generate_password_hash, check_password_hash
from flask import jsonify, request, session,send_file, after_this_request
from flask_restful import Resource, reqparse
import openpyxl, shutil, uuid, os, requests
from werkzeug.utils import secure_filename
from gradetractor import api, jwt, db, app
import pandas as pd

ALLOWED_EXTENSIONS = {'xlsx'}

def allowed_file(filename):

    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def validateNumberOfFiles(path): #function to be used as a wrapper to confirm if both template and db excel fiile has been uploaded

    path = os.path.join(app.config['UPLOAD_FOLDER'], session.get('resultDbPath')) #root path for expcted uploaded document

    totalFiles = [files for files in os.listdir(path) if os.path.isfile(os.path.join(path,files))] #list all the files in the directory

    expectedFiles= ["result.xlsx", "template.xlsx"]

    missingFile = [file for file in expectedFiles if file not in totalFiles] #iterate through to check if files to be worked with are present

    if "template.xlsx" in missingFile:

        return {"error": "Result template not uploaded. Please upload the template file before proceeding."}, 404
        
    if "result.xlsx" in missingFile:

        return {"error": "Results database not uploaded. Please upload the results file before proceeding."}, 404

    return True

@jwt.expired_token_loader
def my_expired_token_callback(jwt_header, jwt_payload):
    return ({"message": "expired token"}), 401

@jwt.invalid_token_loader
def handle_invalid(error):
    return ({"message": "invalid token", "error":error}), 401

class results(Resource):
    @jwt_required()        
    def post(self):

        # Check if the 'file' key exists in the request
        if 'file' not in request.files:

            return {"error": "Results database not uploaded. Please upload the results file before proceeding."}, 400
        
        file = request.files['file']
        
        # Check if the file is of allowed format
        if file and allowed_file(file.filename):

            filename = secure_filename(file.filename)

            # rename file
            filename = "result.xlsx"

            uniqueId = str(uuid.uuid4())

            os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'],uniqueId))

            session["resultDbPath"] = uniqueId

            file.save(os.path.join(app.config['UPLOAD_FOLDER'], uniqueId, filename))

            return {'message': 'File uploaded successfully'}, 200
        
        else:

            return {"error": "Invalid file format. Please upload a file with the .xlsx extension."}, 400
        
class templates(Resource):
    @jwt_required()        
    def post(self):

        if not session.get("resultDbPath"):

            return {"error": "Results database not uploaded. Please upload the results file before proceeding."}, 404
        
        # Check if the 'file' key exists in the request
        if 'file' not in request.files:

            return {"error": "No file uploaded. Please upload a valid file using the 'file' field."}, 400
        
        file = request.files['file']
        
        # Check if the file is of allowed format
        if file and allowed_file(file.filename):

            filename = secure_filename(file.filename)

            # rename file
            filename = "template.xlsx"

            path = session.get("resultDbPath")

            file.save(os.path.join(app.config['UPLOAD_FOLDER'], path, filename))

            return {'message': 'File uploaded successfully'}, 200
        
        else:

            return {"error": "Invalid file format. Please upload a file with the .xlsx extension." }, 400

class allClasses(Resource):
    @jwt_required()
    def get(self):

        if not session.get("resultDbPath"):

            return {"error": "Results database not uploaded. Please upload the results file before proceeding."}, 404
        
        #confirm that all needed file are readily available
        parentPathForFiles = os.path.join(app.config['UPLOAD_FOLDER'], session.get('resultDbPath'))
        
        validator = validateNumberOfFiles(parentPathForFiles)

        if validator != True:

            return validator

        resultSheetPath = os.path.join(app.config['UPLOAD_FOLDER'], session.get('resultDbPath'), 'result.xlsx')

        resultDb = openpyxl.load_workbook(resultSheetPath)
        classes = resultDb.sheetnames

        filteredClasses = []
        unwantedWorkbook = []

        for i in classes:

            if "emy" in i.lower() or "ety" in i.lower():

                filteredClasses.append(i)

        for i in filteredClasses:

            if "Assessment" in i or ">" in i:
                unwantedWorkbook.append(i)

        filteredClasses = [classes for classes in filteredClasses if classes not in unwantedWorkbook]

        return jsonify(allClasses=filteredClasses)

class genResult(Resource):
    @jwt_required()
    def __init__(self):

        self.parser = reqparse.RequestParser()
        self.parser.add_argument("className", required=True)

    def post(self):

        args = self.parser.parse_args()
        className = args["className"]

        if not session.get("resultDbPath"):

            return {"error": "Results database not uploaded. Please upload the results file before proceeding."}, 400
        
        parentPathForFiles = os.path.join(app.config['UPLOAD_FOLDER'], session.get('resultDbPath'))
        
        validator = validateNumberOfFiles(parentPathForFiles)

        if validator != True:

            return validator
        
        #Get all available classes from the allClasses endpoint
        all_classes_url = f"{request.host_url}/api/v1/extractClasses"

        #Extract authorization token to use in making the request
        token = request.headers['Authorization'].split(' ')[1]

        # Include the session cookie in the request
        cookies = request.cookies

        headers = {

            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }

        response = requests.get(all_classes_url, headers=headers, cookies=cookies)
        
        if response.status_code != 200:
            return {"error":response.text}, 500
        
        # Parse the available classes from the response
        available_classes = response.json().get('allClasses', [])

        # Validate if the requested class exists in the available classes
        if className not in available_classes:

            return { "error": f"Class {className} not found. Please ensure the class exists in the uploaded results file." }, 404
                
        resultSheetPath = os.path.join(app.config['UPLOAD_FOLDER'], session.get('resultDbPath'), 'result.xlsx')

        resultDb = openpyxl.load_workbook(resultSheetPath)

        parentPath = os.path.join(app.config['UPLOAD_FOLDER'], session.get('resultDbPath'), className+" report")
        subPath = os.path.join(parentPath, "individual Result")

        os.makedirs(parentPath, exist_ok=True)
        os.makedirs(subPath, exist_ok=True)

        # Input current number of courses
        # selected class that code is working on

        currentClass = resultDb[className]

        # count number of courses 

        # Specify the row number to start the count

        row_number_to_search = 3
        courseCounted = 0
        startColumn = 0

        # get first column that course count begins

        for cell in currentClass[row_number_to_search]:

            cell_value = cell.value

            if cell_value is not None:

                startColumn = cell.column

                break

        # start count to know how many courses

        for cell in currentClass.iter_cols(min_col=startColumn, min_row=row_number_to_search, max_row=row_number_to_search):

            cell_value = cell[0].value

            if cell_value != "AVERAGE":
            
                courseCounted += 1

            else:

                totalCourse = courseCounted + 1 # including column for average(count start from index zero)
                averageValue = courseCounted

                break
            
        # iteration to assign how many student are available while keeping max at 24

        # max number of student in class

        maxStudent = 24 # that is the maximum number of student in a class
        startRow = 7 # from main db, names majorly starts from column 7
        cellCount = 0
        startColumn = 1
        numberOfStudent = 0

        for cell in currentClass.iter_rows(min_col=startColumn, min_row=startRow):
        
            cell_value = cell[0].value

            if cell_value is not None and cellCount <= maxStudent:
            
                numberOfStudent += 1
                cellCount += 1

            else:
            
                break
            
        #course index for db

        courseRowDb = 5
        courseColumnDb = 6

        #course index for template

        courseRowTemplate = 11
        courseColumnTemplate = 3

        # other variable

        scoreList = []
        currentCourseColumn = 11
        trackFailure = 0
        probation = []
        termination = []
        courseNameDb = []
        courseNameTemplate = []

        #load work sheet        

        resultTemplate = os.path.join(app.config['UPLOAD_FOLDER'], session.get('resultDbPath'), 'template.xlsx')
        resultSheetPath = os.path.join(app.config['UPLOAD_FOLDER'], session.get('resultDbPath'), 'result.xlsx')

        template = openpyxl.load_workbook(resultTemplate)
        templateWorksheet = template.worksheets[0]

        data = pd.read_excel(resultSheetPath, sheet_name=className, skiprows=6, header=None, nrows=numberOfStudent) #skip rows is 6 because after that row student names which is the data we need starts counting

        # extract all course name from db and template

        for i in range(courseCounted):
        
            dbCell = currentClass.cell(row=courseRowDb, column=courseColumnDb).value
            courseNameDb.append(dbCell)

            courseColumnDb += 1

        for i in range(courseCounted):
        
            templateCell = templateWorksheet.cell(row=courseRowTemplate, column=courseColumnTemplate).value
            courseNameTemplate.append(templateCell)

            courseRowTemplate += 1

        # iterate through templateWorksheet and manipulate data

        for _, row in data.iterrows():
        
            lastName = row.iloc[1]
            firstName = row.iloc[2]
            middleName = row.iloc[3]

            # Check if any of the names is NaN, if yes, set them to empty strings

            if pd.isna(firstName):
            
                firstName = ""

            if pd.isna(middleName):
            
                middleName = ""

            # ommited last name(every individual would surely have a last name) to keep track of empty result that would be called nan

            score = row.iloc[5:].values

            score = pd.Series(score).fillna('NA').values

            for scores in score[:totalCourse]:
            
                scoreList.append(scores)

            average = scoreList[averageValue]

            # check for failures

            for scores in scoreList[:averageValue]:
            
                try:
                
                    if int(scores) < 60:
                    
                        trackFailure += 1

                except:
                
                    pass
                
            # Create a dictionary to store the mapping of course names and their corresponding scores
            course_score_mapping = dict(zip(courseNameDb, score[:totalCourse]))

            # Iterate through the course names in the template
            for coursesTemplate in courseNameTemplate:
            
                # Check if the course name exists in the mapping
                if coursesTemplate in course_score_mapping:
                
                    # Get the score from the mapping for the current course name
                    score_for_course = course_score_mapping[coursesTemplate]

                else:
                
                    # If the course name is not found in the mapping, set the score to 'NA'
                    score_for_course = ' '
                # Write the score to the corresponding cell in the template

                templateWorksheet.cell(row=currentCourseColumn, column=4, value=score_for_course)
                currentCourseColumn += 1

            name = str(lastName) + " " + str(firstName) + " " + str(middleName)

            templateWorksheet.cell(row=7, column=3, value=name)
            templateWorksheet.cell(row=7, column=5, value=average)
            templateWorksheet.cell(row=5, column=5, value=className)

            # check number of failed courses

            if trackFailure == 2:
            
                probation.append(name)

            if trackFailure > 2:
            
                termination.append(name)

            # clear variable memory

            trackFailure = 0
            scoreList = []
            currentCourseColumn = 11

            #save file

            try:
            
                template.save('{}/{}.xlsx'.format(subPath,name))

            except:
            
                pass
            
        # loop to delete null report file created cause of incomplete class

        for filename in os.listdir(subPath):

            if "nan" in filename:
            
                file_path = os.path.join(subPath, filename)

                os.remove(file_path)

        # create file to store probation and termination list

        probationFilePath = os.path.join(parentPath, className + " Probation List.txt")
        terminationFilePath = os.path.join(parentPath, className + " Termination List.txt")

        probationFile = open(probationFilePath, "w")
        terminationFile = open(terminationFilePath, "w")

        # generate student on probation

        for watchlist in probation:

            probationFile.write(watchlist + '\n')

        # generate student on termination

        for watchlist in termination:

            terminationFile.write(watchlist + '\n')

        # Save and close files before sending

        probationFile.close()
        terminationFile.close()

        # zip folder containing results

        resultPath = os.path.join(app.config['UPLOAD_FOLDER'], session.get('resultDbPath'), className + " generated result")
        zippedResultPath = shutil.make_archive(resultPath, 'zip', parentPath)

        @after_this_request
        def deleteAllFiles(response):

            userDirectoryToDelete = os.path.join(app.config['UPLOAD_FOLDER'], session.get('resultDbPath'))
                                      
            shutil.rmtree(userDirectoryToDelete)
            session.pop("resultDbPath", None)
            return response

        return send_file(zippedResultPath, as_attachment=True, mimetype='application/zip')

class login(Resource):

    def __init__(self):

        self.parser = reqparse.RequestParser()
        self.parser.add_argument("email", required=True)
        self.parser.add_argument("password", required=True)
    
    def post(self):

        args = self.parser.parse_args()
        email = args["email"].lower()
        password = args["password"]

        checkUser = db.users.find_one({"email":email})

        if not checkUser or not check_password_hash(checkUser.get("password"), password):

            return ({"Error":"Email or password is incorrect"})
        

        #To retrive the email use sub as the key for the query

        refresh_token = create_refresh_token(identity=email)
        access_token = create_access_token(identity=email, fresh=True, additional_claims={"refresh_jti": decode_token(refresh_token)["jti"]})

        response = jsonify(

                {
                    "success": "login successful",
                    "token":{
                        "access_token":access_token, "refresh_token":refresh_token
                    }
                }
                
            )

        set_access_cookies(response, access_token)
            
        return response
        
class reg(Resource):

    def __init__(self):

        self.parser = reqparse.RequestParser()
        self.parser.add_argument("email", required=True)
        self.parser.add_argument("password", required=True)
    
    def post(self):

        args = self.parser.parse_args()
        email = args["email"].lower()
        password = args["password"]

        checkExistingMail = db.users.find_one({"email":email})

        if checkExistingMail:

            return ({"Error":"Mail already exist"})
        
        hashedPass = generate_password_hash(password)

        data = {'email':email, 'password':hashedPass}
        result = db.users.insert_one(data)

        if result.inserted_id:
            
            return ({"message":"Data inserted successfully", "inserted_id":str(result.inserted_id)}), 201
        
        else:

            return jsonify({"error": "Failed to save your data. Please try again later. If the issue persists, contact support."}), 500

api.add_resource(reg, '/api/v1/reg', '/api/v1/reg/')
api.add_resource(login, '/api/v1/login', '/api/v1/login/')
api.add_resource(results, '/api/v1/result', '/api/v1/result/')
api.add_resource(allClasses, '/api/v1/extractClasses', '/api/v1/extractClasses/')
api.add_resource(templates, '/api/v1/template', '/api/v1/template/')
api.add_resource(genResult, '/api/v1/genResult', '/api/v1/genResult/')